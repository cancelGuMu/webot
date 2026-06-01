"""Reconcile WeFlow wxids with exported chat display names.

Reads:
  - data/messages.db
  - data/nicknames.json
  - wechatdata/*.xlsx

Writes:
  - data/nicknames.json
  - data/nickname_reconcile_report.json
"""

from __future__ import annotations

import json
import re
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "data" / "messages.db"
NICKNAMES_PATH = ROOT / "data" / "nicknames.json"
REPORT_PATH = ROOT / "data" / "nickname_reconcile_report.json"
XLSX_DIR = ROOT / "wechatdata"

MAX_TIME_DELTA_SEC = 180
MIN_VOTES = 2
MIN_SHARE = 0.67


def normalize_content(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\u2005", " ").replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text.strip())
    aliases = {
        "[表情包]": "[表情]",
        "[动画表情]": "[表情]",
    }
    return aliases.get(text, text)


def parse_excel_time(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def read_db_messages() -> list[dict[str, Any]]:
    uri = f"file:{DB_PATH.as_posix()}?mode=ro"
    conn = sqlite3.connect(uri, uri=True, timeout=2)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """
        SELECT id, chat_id, sender_id, sender_name, content, timestamp
        FROM messages
        WHERE chat_id LIKE '%@chatroom'
        ORDER BY timestamp ASC
        """
    ).fetchall()
    conn.close()
    return [
        {
            **dict(row),
            "dt": datetime.fromtimestamp(int(row["timestamp"])),
            "norm_content": normalize_content(row["content"]),
        }
        for row in rows
    ]


def read_excel_messages() -> list[dict[str, Any]]:
    files = sorted(XLSX_DIR.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No .xlsx files found in {XLSX_DIR}")

    xlsx_path = files[0]
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    messages: list[dict[str, Any]] = []

    for row in sheet.iter_rows(min_row=5, values_only=True):
        seq, time_value, sender, msg_type, content, *_ = row
        if seq is None:
            continue
        dt = parse_excel_time(time_value)
        if dt is None:
            continue
        messages.append(
            {
                "seq": seq,
                "dt": dt,
                "sender": "" if sender is None else str(sender).strip(),
                "msg_type": "" if msg_type is None else str(msg_type).strip(),
                "content": "" if content is None else str(content),
                "norm_content": normalize_content(content),
            }
        )

    return messages


def index_excel(messages: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    indexed: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for msg in messages:
        if msg["norm_content"]:
            indexed[msg["norm_content"]].append(msg)
    return indexed


def best_match(
    db_msg: dict[str, Any],
    candidates: list[dict[str, Any]],
) -> dict[str, Any] | None:
    near = []
    for candidate in candidates:
        delta = abs((candidate["dt"] - db_msg["dt"]).total_seconds())
        if delta <= MAX_TIME_DELTA_SEC:
            near.append((delta, candidate))
    if not near:
        return None
    near.sort(key=lambda item: item[0])
    if len(near) > 1 and near[0][0] == near[1][0] and near[0][1]["sender"] != near[1][1]["sender"]:
        return None
    return {**near[0][1], "delta_sec": near[0][0]}


def load_existing_nicknames() -> dict[str, str]:
    if not NICKNAMES_PATH.exists():
        return {}
    with NICKNAMES_PATH.open("r", encoding="utf-8") as f:
        data = json.load(f)
    return {str(k): str(v) for k, v in data.items()}


def main() -> None:
    db_messages = read_db_messages()
    excel_messages = read_excel_messages()
    excel_index = index_excel(excel_messages)
    existing = load_existing_nicknames()

    votes: dict[str, Counter[str]] = defaultdict(Counter)
    examples: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    matched_rows = 0

    for db_msg in db_messages:
        content = db_msg["norm_content"]
        if not content:
            continue
        match = best_match(db_msg, excel_index.get(content, []))
        if match is None:
            continue
        sender = match["sender"]
        if not sender:
            continue
        matched_rows += 1
        wxid = db_msg["sender_id"]
        votes[wxid][sender] += 1
        if len(examples[wxid][sender]) < 5:
            examples[wxid][sender].append(
                {
                    "db_id": db_msg["id"],
                    "db_time": db_msg["dt"].strftime("%Y-%m-%d %H:%M:%S"),
                    "excel_seq": match["seq"],
                    "excel_time": match["dt"].strftime("%Y-%m-%d %H:%M:%S"),
                    "delta_sec": match["delta_sec"],
                    "content": db_msg["norm_content"],
                }
            )

    updates: dict[str, str] = {}
    decisions = []
    for wxid, counter in sorted(votes.items()):
        total = sum(counter.values())
        name, count = counter.most_common(1)[0]
        share = count / total if total else 0
        existing_name = existing.get(wxid, "")
        accepted = count >= MIN_VOTES and share >= MIN_SHARE

        decision = {
            "wxid": wxid,
            "existing": existing_name,
            "candidate": name,
            "accepted": accepted,
            "votes": dict(counter.most_common()),
            "top_votes": count,
            "total_votes": total,
            "share": round(share, 3),
            "examples": examples[wxid][name],
        }
        decisions.append(decision)

        if accepted and name:
            updates[wxid] = name

    merged = dict(existing)
    for wxid, name in updates.items():
        if wxid == "_comment":
            continue
        merged[wxid] = name

    # Keep comment first for readability.
    ordered: dict[str, str] = {}
    if "_comment" in merged:
        ordered["_comment"] = merged.pop("_comment")
    for key in sorted(merged):
        ordered[key] = merged[key]

    with NICKNAMES_PATH.open("w", encoding="utf-8") as f:
        json.dump(ordered, f, ensure_ascii=False, indent=2)
        f.write("\n")

    report = {
        "db_message_count": len(db_messages),
        "excel_message_count": len(excel_messages),
        "matched_message_count": matched_rows,
        "accepted_update_count": len(updates),
        "updates": updates,
        "decisions": decisions,
    }
    with REPORT_PATH.open("w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(json.dumps({
        "db_message_count": len(db_messages),
        "excel_message_count": len(excel_messages),
        "matched_message_count": matched_rows,
        "accepted_update_count": len(updates),
        "updates": updates,
        "report": str(REPORT_PATH),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
